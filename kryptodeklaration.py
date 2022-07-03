#!/usr/bin/env python3
#
# kryptodeklaration.py, Thomas Lundqvist, 2020-2022
#
# Räkna ut vinst, förlust och utgående genomsnittligt omkostnadsbelopp
# 

import sys, datetime
from collections import OrderedDict
import openpyxl

# Inflikar:
SHEET_TRAN = "Transaktioner"
SHEET_INBAL = "Inbalans"

# Utflikar:
SHEET_RESULTAT = "Resultat"
SHEET_UTBAL = "Utbalans"

DIV = "*" * 60

class Konto:
    """Håller innehav och genomsnittligt omkostnadsbelopp för en kryptovaluta"""
    def __init__(self, namn, enhet, innehav, gob):
        self.namn = namn
        self.enhet = enhet
        self.innehav = innehav
        self.gob = gob
        self._totbelopp = innehav * gob
        # Summeringar för deklaration, dela upp vinst och förlust på två poster
        # Ränteinkomster redovisas separat
        self._dekl_vinst_sälj = 0
        self._dekl_vinst_sälj_belopp = 0
        self._dekl_vinst_omkostnad = 0
        self._dekl_vinst = 0
        self._dekl_förlust_sälj = 0
        self._dekl_förlust_sälj_belopp = 0
        self._dekl_förlust_omkostnad = 0
        self._dekl_förlust = 0
        self._dekl_ränta = 0

    def getAll(self):
        return [self.namn, self.enhet, self.innehav, self.gob]
    
    def update(self, datum, händelse, antal, belopp):
        vinst = None
        omkostnad = None
        ränta = None
        if händelse == "köp":
            if antal < 0:
                print("Varning, antal mindre än noll vid köp.", self.enhet, datum)
            self._totbelopp += belopp
            self.innehav += antal
            self.gob = self._totbelopp / self.innehav
        elif händelse == "sälj":
            if antal > 0:
                print("Varning, antal positivt vid sälj! Säljantal ska vara negativt.",
                      self.enhet, datum)
            self.innehav += antal   # antal redan < 0 vid sälj
            if self.innehav < 0:
                print("Innehav:", self.innehav, "efter", händelse, antal, belopp)
                sys.exit("Error: innehav < 0 för " + self.enhet, datum)
            omkostnad = -antal * self.gob
            vinst = belopp - omkostnad
            self._totbelopp = self.innehav * self.gob
            if vinst >= 0:
                self._dekl_vinst_sälj += antal
                self._dekl_vinst_sälj_belopp += belopp
                self._dekl_vinst_omkostnad += omkostnad
                self._dekl_vinst += vinst
            else:
                self._dekl_förlust_sälj += antal
                self._dekl_förlust_sälj_belopp += belopp
                self._dekl_förlust_omkostnad += omkostnad
                self._dekl_förlust += vinst
            if self.innehav == 0:
                self._totbelopp = 0
                self.gob = 0
        elif händelse == "ränta":
            # Ränta betraktas som köp till aktuell kurs samtidigt som samma
            # belopp ska bokföras som ränteinkomst
            if antal < 0 or belopp < 0:
                print("Varning: antal eller belopp negativt. Båda bör vara positiva vid ränta!",
                      self.enhet, datum)
            # Ränta ger allt som ränteinkomst direkt
            ränta = belopp
            self._dekl_ränta += ränta
            # Därefter samma som köp
            self._totbelopp += belopp
            self.innehav += antal
            self.gob = self._totbelopp / self.innehav
        else:
            sys.exit("Error: okänd händelse i transaktion: " + händelse + " " + datum)
        return omkostnad, vinst, ränta

    def get_dekl_vinst(self):
        return (self._dekl_vinst_sälj,
                self._dekl_vinst_sälj_belopp,
                self._dekl_vinst_omkostnad,
                self._dekl_vinst)
    def get_dekl_förlust(self):
        return (self._dekl_förlust_sälj,
                self._dekl_förlust_sälj_belopp,
                self._dekl_förlust_omkostnad,
                self._dekl_förlust)

    def __lt__(self, other):
        """Reverse sort på totalt omkostbelopp, om lika normal sort på namn"""
        b1 = self._totbelopp
        b2 = other._totbelopp
        return self.namn < other.namn if b1 == b2 else b1 > b2
    
    def __repr__(self):
        """Vänlig utskrift för print()"""
        return 'Konto("%s","%s","%s","%s")' % (self.namn, self.enhet, self.innehav, self.gob)

class Transaktion:
    """Håller en transaktionsrad, "köp", "sälj" eller "ränta" som händelse"""
    def __init__(self, datum, var, händelse, antal, valuta, belopp):
        self.datum = datum
        self.var = var
        self.händelse = händelse
        self.antal = antal
        self.valuta = valuta
        self.belopp = belopp
        
    def getAll(self):
        return [self.datum, self.var, self.händelse, self.antal, self.valuta, self.belopp]
    
    def __lt__(self, other):
        """Sort på datum"""
        return self.datum < other.datum
        
    def __repr__(self):
        """Vänlig utskrift för print()"""
        return 'Transaktion("%s","%s","%s","%s","%s","%s")' % (
            self.datum, self.var, self.händelse,
            self.antal, self.valuta, self.belopp)

class Kalkfil():
    def __init__(self):
        errortext = """Användning:
  -  Ange en indata-excelfil som argument (xlsx-fil)
  -  Excelfilen måste ha flikarna "Transaktioner" och "Inbalans"
  -  Utdata skapas i samma excelfil i två nya flikar: "Resultat"
     och "Utbalans".
        """
        if len(sys.argv) < 2:
            sys.exit(errortext)

        print("Läser och modifierar", sys.argv[1])
        print(DIV)
        try:
            self.filename = sys.argv[1]
            workbook = openpyxl.load_workbook(filename = self.filename)
            print("Befintliga flikar:")
            for s in workbook.sheetnames:
                print("  ", s)
            sheet_tran = workbook[SHEET_TRAN]
            sheet_inbal = workbook[SHEET_INBAL]
        except FileNotFoundError:
            sys.exit("Error: File not found!")
        except KeyError:
            sys.exit("Error: Hittar ej rätt flikar!")

        print("Skriver resultat till flikarna:")
        print("  ", SHEET_RESULTAT, "(OBS, finns redan)" if SHEET_RESULTAT in workbook else "")
        print("  ", SHEET_UTBAL, "(OBS, finns redan)" if SHEET_UTBAL in workbook else "")
        print(DIV)
        if SHEET_RESULTAT in workbook or SHEET_UTBAL in workbook:
            print("Varning, någon av utflikarna finns redan och kommer att ersättas.")
            i = input("Fortsätta? (j/n) ")
            if not i in 'jJyY':
                sys.exit("Avbryter!")
            print(DIV)
        if SHEET_RESULTAT in workbook:
            workbook.remove(workbook[SHEET_RESULTAT])
        if SHEET_UTBAL in workbook:
            workbook.remove(workbook[SHEET_UTBAL])
        sheetR = workbook.create_sheet(SHEET_RESULTAT)
        sheetU = workbook.create_sheet(SHEET_UTBAL)

        self.sheetT, self.sheetI = sheet_tran, sheet_inbal
        self.sheetR, self.sheetU = sheetR, sheetU
        self.workbook = workbook

    def save(self):
        self.workbook.save(self.filename)
    
# Läser in fliken "Inbalans"
#
# Returnerar en OrderedDict med konton. Exempel:
#   { 'GBYTE': Konto("Obyte","GBYTE","0.822","2017.031630"),
#     'mETH': Konto("Ethereum","mETH","1493.19","1.6826016"),
#     'ADA': Konto("Cardano","ADA","1377.0","0.0"),
#     'LTC': Konto("Litecoin","LTC","3.124","0.0") }
# med kontona i samma ordning som i Inbalans

def read_inbalans(sheet):
    balans = OrderedDict()

    foundtable = False
    for row in sheet.iter_rows(values_only = True):
        if foundtable:
            try:
                namn = row[col_namn]
                enhet = row[col_enhet]
                innehav = float(row[col_innehav])
                gob = float(row[col_gob])
#                print(namn, enhet, innehav, gob)
                konto = Konto(namn, enhet, innehav, gob)
                balans[enhet] = konto
            except TypeError:
                pass
        else:
            try:
                col_namn = row.index("Namn")
                col_enhet = row.index("Enhet")
                col_innehav = row.index("Innehav")
                col_gob = row.index("GOB")
#                print("Hittade inbalanstabellen")
                foundtable = True
            except ValueError:
                pass
    print("Läst Inbalans:", len(balans), "valutor")
    return balans

# Läser alla transaktionsrader från fliken "Transaktioner"
#
# Returnerar lista med transaktionsrader. Exempel:
#   [ Transaktion("2019-01-10 00:00:00","coinbase","köp","50.02232","mBTC","1712.0"),
#     Transaktion("2019-01-11 00:00:00","bittrex","sälj","-80.0","mBTC","2611.0"),
#     Transaktion("2019-01-11 00:00:00","bittrex","köp","30.0","REP","2611.0") ]

def read_transactions(sheet):
    translist = []
    foundtable = False
    for row in sheet.iter_rows(values_only = True):
        if foundtable:
            try:
                datum = row[col_datum]
                var = row[col_var]
                händelse = row[col_händelse]
                antal = float(row[col_antal])
                valuta = row[col_valuta]
                belopp = float(row[col_belopp])
#                print(datum, var, händelse, antal, valuta, belopp)
                # Om datum och var är tomma, kopiera från föregående rad
                if not datum or str(datum).strip() == "":
                    datum = old_datum
                if not var or var.strip() == "":
                    var = old_var
                if type(datum) == str:
                    datum = datetime.datetime.strptime(datum, "%Y-%m-%d")
#                print(datum, var, händelse, antal, valuta, belopp)
                old_datum = datum
                old_var = var
                # Omvandling enheter, gillar milli BTC/ETH bättre
                if valuta == "BTC":
                    valuta = "mBTC"
                    antal *= 1000
                if valuta == "ETH":
                    valuta = "mETH"
                    antal *= 1000
                trans = Transaktion(datum, var, händelse, antal, valuta, belopp)
                translist.append(trans)
            except (TypeError, ValueError):
                pass
        else:
            try:
                col_datum = row.index("Datum")
                col_var = row.index("Var")
                col_händelse = row.index("Händelse")
                col_antal = row.index("Antal")
                col_valuta = row.index("Valuta")
                col_belopp = row.index("Belopp")
#                print("Hittade transaktionstabellen")
                foundtable = True
            except ValueError:
                pass
    print("Läst Transaktioner:", len(translist), "rader")
    return translist

# Sprid ut transaktionsraderna sorterade på valutan i en dict.
# Kontrollerar också så att alla valutor i transaktionerna finns med i inbalans.
#
# Returnerar dict med uppdelade transaktionsrader. Exempel:
#   {'mBTC': [ Transaktion("2019-01-10 00:00:00","coinbase","köp","50.02232","mBTC","1712.0"),
#     Transaktion("2019-01-11 00:00:00","bittrex","sälj","-80.0","mBTC","2611.0") ],
#    'REP': [ Transaktion("2019-01-11 00:00:00","bittrex","köp","30.0","REP","2611.0") ] }

def sort_check_transactions(balans, translist):
    transtable = {}
    for tx in translist:
        transtable.setdefault(tx.valuta, []).append(tx)
    for valuta in transtable.keys():
        transtable[valuta].sort()
    diff = transtable.keys() - balans.keys()   # set difference
    if len(diff) > 0:
        sys.exit("Error, några valutor saknas i inbalansen:\n" + str(diff))
    return transtable

# Skriv ut resultatfliken "Resultat"
#
# Skapar många små tabeller för var sin valuta med
# extra kolumner för vinst, förlust och ränteberäkningarna.
#
# Uppdaterar löpande alla konton i "balans"-dicten.

ROWHEIGHT = 12.8 # verkar lagom stort

def output_results(sheet, balans, transtable):
    boldfont = openpyxl.styles.Font(name='Arial',size=10,bold=True)
    sheet["A1"].value = "Resultat"
    sheet["A1"].font = boldfont
    for i in range(3):
        sheet.row_dimensions[i+1].height = ROWHEIGHT
    row = 4
    # Utgå från valutorna i inbalansen för att få samma sorteringsordning
    # Dessa är ett superset av tx-valutorna
    valutor = balans.keys()
    h = ["Datum","Var","Händelse","Antal+","Antal-","Valuta",
         "Belopp+", "Belopp-", None,
         "Innehav", "GOB", "Omkostnad", "Vinst", "Förlust", "Ränta"]
    tot_vinst = 0
    tot_förlust = 0
    tot_ränta = 0
    for valuta in valutor:
        if not valuta in transtable.keys():
            # Alla inbalansvalutor finns kanske inte som transaktioner
            continue
        newh = [valuta] + h
        for c,v in enumerate(newh):
            cell = sheet.cell(row=row, column=c+1)
            cell.value = v
            cell.font = boldfont
            if v not in [valuta, "Datum", "Var", "Händelse", "Valuta"]:
                cell.alignment = openpyxl.styles.Alignment(horizontal="right")
        sheet.row_dimensions[row].height = ROWHEIGHT
        row += 1
        try:
            konto = balans[valuta]
        except KeyError:
            sys.exit("Error: " + valuta + " finns inte i inbalans")
        sheet.cell(row=row, column=11).value = konto.innehav
        sheet.cell(row=row, column=12).value = konto.gob
        sheet.row_dimensions[row].height = ROWHEIGHT
        row += 1
        dekl_sälj = 0
        dekl_sälj_belopp = 0;
        for tx in transtable[valuta]:
            v = tx.getAll()
            sheet.cell(row=row, column=2).value = v[0] # Datum
            sheet.cell(row=row, column=2).alignment = openpyxl.styles.Alignment(horizontal="left")
            sheet.cell(row=row, column=3).value = v[1] # Var
            sheet.cell(row=row, column=4).value = v[2] # Händelse
            if v[3] > 0:
                sheet.cell(row=row, column=5).value = v[3] # Debet+
                sheet.cell(row=row, column=8).value = v[5] # Belopp+
            else:
                sheet.cell(row=row, column=6).value = v[3] # Kredit-
                sheet.cell(row=row, column=9).value = v[5] # Belopp-
                dekl_sälj += v[3]
                dekl_sälj_belopp += v[5]
            sheet.cell(row=row, column=7).value = v[4] # Valuta
            sheet.cell(row=row, column=2).number_format = 'YYYY-MM-DD'
            omkostnad, vinst, ränta = konto.update(tx.datum, tx.händelse, tx.antal, tx.belopp)
            sheet.cell(row=row, column=11).value = konto.innehav
            sheet.cell(row=row, column=12).value = konto.gob
            if vinst != None:
                sheet.cell(row=row, column=13).value = omkostnad
                if vinst >= 0:
                    col = 14
                    tot_vinst += vinst
                else:
                    col = 15
                    tot_förlust += vinst
                sheet.cell(row=row, column=col).value = vinst
            if ränta != None:
                sheet.cell(row=row, column=16).value = ränta
                tot_ränta += ränta
            sheet.row_dimensions[row].height = ROWHEIGHT
            row += 1
        dsälj, dbelopp, domkostnad, dvinst = konto.get_dekl_vinst()
        if dvinst > 0:
            sheet.cell(row=row, column=3).value = "Deklaration vinst"
            sheet.cell(row=row, column=6).value = dsälj
            sheet.cell(row=row, column=9).value = dbelopp
            sheet.cell(row=row, column=13).value = domkostnad
            sheet.cell(row=row, column=14).value = dvinst
            for i in [3, 6, 9, 13, 14]:
                sheet.cell(row=row, column=i).font = boldfont
                if i >= 9:
                    sheet.cell(row=row, column=i).number_format = "0.00"
            row += 1
        dsälj, dbelopp, domkostnad, dvinst = konto.get_dekl_förlust()
        if dvinst < 0:
            sheet.cell(row=row, column=3).value = "Deklaration förlust"
            sheet.cell(row=row, column=6).value = dsälj
            sheet.cell(row=row, column=9).value = dbelopp
            sheet.cell(row=row, column=13).value = domkostnad
            sheet.cell(row=row, column=15).value = dvinst
            for i in [3, 6, 9, 13, 14, 15]:
                sheet.cell(row=row, column=i).font = boldfont
                if i >= 9:
                    sheet.cell(row=row, column=i).number_format = "0.00"
            row += 1
        row += 1
    row += 1
    sheet.cell(row=row, column=14).value = "TOTALT"
    sheet.cell(row=row, column=14).font = boldfont
    for c in range(3):
        sheet.cell(row=row+1, column=14+c).value = ["Vinst", "Förlust", "Ränta"][c]
        sheet.cell(row=row+1, column=14+c).font = boldfont
        sheet.cell(row=row+2, column=14+c).value = [tot_vinst, tot_förlust, tot_ränta][c]
    for i in range(3):
        sheet.row_dimensions[row+i].height = ROWHEIGHT
        
    row += 4
    skatt = (tot_vinst + tot_ränta + tot_förlust*0.7)*0.3
    sheet.cell(row=row, column=14).value = "SKATT"
    sheet.cell(row=row, column=14).font = boldfont
    sheet.cell(row=row+1, column=14).value = skatt
    
    # Set fairly sensible column widths. Width is expressed as the
    # number of monospace characters. Will do even for other fonts.
    sheet.column_dimensions["A"].width = 14
    sheet.column_dimensions["B"].width = 11
    sheet.column_dimensions["C"].width = 20
    sheet.column_dimensions["G"].width = 14
    for c in "HIMNOP":
        sheet.column_dimensions[c].width = 11
        sheet.column_dimensions[c].number_format = "0.00"
    

    print("Skapat ny flik", SHEET_RESULTAT)
    print("Total vinst:  ", tot_vinst)
    print("Total förlust:", tot_förlust)
    print("Total ränta:  ", tot_ränta)
    print("Total skatt:  ", skatt)

# Skriv ut utbalansfliken. Måste köras sist då balansen är uppdaterad med
# alla transaktioner.

def output_utbalans(sheet, balans):
    konton = list(balans.values())
    konton.sort()
#    print(konton)
    boldfont = openpyxl.styles.Font(name='Arial',size=10,bold=True)
    sheet["A1"].value = "Utgående balans"
    sheet["A1"].font = boldfont
    for i in range(3):
        sheet.row_dimensions[i+1].height = ROWHEIGHT
    h = ["Namn","Enhet","Innehav","GOB","","Omkostnad"]
    for c in range(1,len(h)+1):
        cell = sheet.cell(row=3, column=c)
        cell.value = h[c-1]
        cell.font = boldfont
    row = 4
    for k in konton:
        for i,v in enumerate(k.getAll()):
            sheet.cell(row=row, column=i+1).value = v
        sheet.cell(row=row, column=6).value = k.innehav * k.gob
        sheet.row_dimensions[row].height = ROWHEIGHT
        row += 1
    # Set fairly sensible column widths. Width is expressed as the
    # number of monospace characters. Will do even for other fonts.
    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 15
    sheet.column_dimensions["F"].number_format = "0"
    print("Skapat ny flik", SHEET_UTBAL)


kalkfil = Kalkfil()
balans = read_inbalans(kalkfil.sheetI)
translist = read_transactions(kalkfil.sheetT)
transtable = sort_check_transactions(balans, translist)
output_results(kalkfil.sheetR, balans, transtable)
output_utbalans(kalkfil.sheetU, balans)
kalkfil.save()
print(DIV)
print("Klar!")

#print(balans)
#print(transtable)
